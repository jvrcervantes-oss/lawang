#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_plan.py — Generador de planes financieros LAWANG (Lamborghini Wild Coast Sumba).

Construye un Excel corporativo de marca DESDE CERO (no usa la plantilla del cliente):
informe editorial legible, paleta Lawang, portada de marca incrustada.

El JSON de datos es la única fuente de verdad. Para actualizar cifras -> editar el JSON
y regenerar. El Excel es el documento renderizado.

Uso:
    python build_plan.py inbox/datos.json
    python build_plan.py inbox/datos.json salida/Plan.xlsx
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Falta openpyxl. Instala con:  pip install openpyxl")

BASE = Path(__file__).resolve().parent

# ----------------------------------------------------------------------------- PALETA LAWANG
LINEN    = "FFF5F0E6"   # Raw Linen  · fondo 60%
LINEN_LT = "FFFBF8F2"   # banding claro
GREEN    = "FF485B37"   # Territorial Green · cabeceras
LAGOON   = "FF104C4F"   # Deep Lagoon · totales / KPI
ASH      = "FF2E3437"   # Volcanic Ash · texto
CANOPY   = "FF8F9B7A"   # Soft Canopy · acento
SAND     = "FFBEB3A5"   # Stone Sand · reglas / neutro

EUR = '"€"#,##0;[Red]("€"#,##0)'
PCT = '0.0%'

rule       = Side(style="thin",   color=SAND)
rule_med   = Side(style="medium", color=SAND)
B_BOTTOM   = Border(bottom=rule)
B_TOP_MED  = Border(top=rule_med)

def F(size, bold=False, color=ASH, name="Segoe UI", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

FILL = lambda c: PatternFill("solid", fgColor=c)
RIGHT  = Alignment(horizontal="right",  vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", indent=1)
LEFT0  = Alignment(horizontal="left",   vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ----------------------------------------------------------------------------- MODELO
def compute(d):
    g = lambda k, default=0: d.get(k) if d.get(k) is not None else default
    villas   = g("num_villas")
    precio   = g("precio_villa")
    cconstr  = g("coste_construccion_villa")
    gross    = villas * precio
    construc = villas * cconstr
    suelo    = g("coste_adquisicion_suelo")
    urb      = g("urbanizacion")
    lic      = g("licencias")
    direct   = construc + suelo + urb + lic
    nick     = construc * g("pct_nick_maltese")
    arq      = g("arquitecto_local"); otros = g("otros_tecnicos")
    prof     = nick + arq + otros
    mkt      = g("marketing")
    comis    = gross * g("pct_comisiones")
    legal    = g("legal_notaria"); adm = g("gastos_admin")
    seg      = gross * g("pct_seguros")
    comadm   = mkt + comis + legal + adm + seg
    total    = direct + prof + comadm
    op       = gross - total
    tonino   = op * g("pct_tonino"); intro = op * g("pct_introductor")
    net      = op - tonino - intro
    inv      = gross * g("pct_aportacion_socios") + gross * g("pct_deuda")
    land_pot = g("suelo_total") * g("precio_suelo_m2")
    return dict(
        villas=villas, precio=precio, gross=gross, construc=construc, suelo=suelo,
        urb=urb, lic=lic, direct=direct, nick=nick, arq=arq, otros=otros, prof=prof,
        mkt=mkt, comis=comis, legal=legal, adm=adm, seg=seg, comadm=comadm,
        total=total, op=op, tonino=tonino, intro=intro, net=net,
        net_margin=(net / gross if gross else 0),
        roi_cost=(net / total if total else 0),
        roi_inv=(net / inv if inv else 0),
        land_pot=land_pot, land_upside=land_pot - suelo,
        pct=lambda k: d.get(k, 0),
    )


# ----------------------------------------------------------------------------- SUPUESTOS
# Orden = filas B2..B23 de la hoja "Supuestos". (key, etiqueta, unidad, es_pct)
SUP_FIELDS = [
    ("num_villas",               "Nº villas (fase)",                          "villas", False),
    ("precio_villa",             "Precio de venta por villa",                 "€",      False),
    ("superficie_villa",         "Superficie por villa",                      "m²",     False),
    ("coste_construccion_villa", "Coste de construcción por villa",           "€",      False),
    ("suelo_total",              "Superficie total del proyecto",             "m²",     False),
    ("precio_suelo_m2",          "Precio / valor objetivo del suelo",         "€/m²",   False),
    ("coste_adquisicion_suelo",  "Coste de adquisición del suelo",            "€",      False),
    ("urbanizacion",             "Urbanización & infraestructura",            "€",      False),
    ("licencias",                "Licencias & permisos",                      "€",      False),
    ("pct_nick_maltese",         "Nick Maltese Studio (% s/ construcción)",   "%",      True),
    ("arquitecto_local",         "Arquitecto local & ingeniería",             "€",      False),
    ("otros_tecnicos",           "Otros profesionales técnicos",              "€",      False),
    ("marketing",                "Marketing & branding",                      "€",      False),
    ("pct_comisiones",           "Comisiones comerciales (% s/ ventas)",      "%",      True),
    ("legal_notaria",            "Gastos legales & notaría",                  "€",      False),
    ("gastos_admin",             "Gastos administrativos",                    "€",      False),
    ("pct_seguros",              "Seguros & contingencias (% s/ ventas)",     "%",      True),
    ("pct_tonino",               "Tonino Lamborghini (% s/ beneficio)",       "%",      True),
    ("pct_introductor",          "Introductor de marca (% s/ beneficio)",     "%",      True),
    ("pct_preventas",            "Preventas (% s/ ventas)",                   "%",      True),
    ("pct_aportacion_socios",    "Aportación socios (% s/ ventas)",           "%",      True),
    ("pct_deuda",                "Deuda / crédito puente (% s/ ventas)",      "%",      True),
]
SUP = {key: f"B{i+2}" for i, (key, *_ ) in enumerate(SUP_FIELDS)}


def build_supuestos(wb, d):
    """Hoja editable con los inputs. El informe los referencia por fórmula."""
    ws = wb.create_sheet("Supuestos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    ws.merge_cells("B2:D2")
    h = ws["B2"]; h.value = "SUPUESTOS — EDITABLES"
    h.font = F(13, True, LINEN, "Georgia"); h.alignment = LEFT
    for col in "BCD": ws[f"{col}2"].fill = FILL(GREEN)
    ws.row_dimensions[2].height = 26
    sub = ws["B3"]; sub.value = "Cambia un valor y todo el Plan Financiero se recalcula."
    sub.font = F(9, False, SAND, italic=True); sub.alignment = LEFT
    ws.row_dimensions[3].height = 16

    g = lambda k: d.get(k) if d.get(k) is not None else 0
    start = 5  # los inputs empiezan en la fila 5 (C5..C26), bajo la cabecera
    global SUP
    SUP = {key: f"C{start + i}" for i, (key, *_ ) in enumerate(SUP_FIELDS)}
    for i, (key, label, unit, is_pct) in enumerate(SUP_FIELDS):
        r = start + i
        lc = ws[f"B{r}"]; lc.value = label
        lc.font = F(11, False, ASH); lc.alignment = LEFT
        vc = ws[f"C{r}"]; vc.value = g(key)
        vc.font = F(11, True, LAGOON); vc.alignment = RIGHT
        vc.number_format = PCT if is_pct else (EUR if unit in ("€", "€/m²") else "#,##0")
        uc = ws[f"D{r}"]; uc.value = unit
        uc.font = F(9, False, SAND); uc.alignment = LEFT0
        for col in "BCD":
            ws[f"{col}{r}"].border = B_BOTTOM
            ws[f"{col}{r}"].fill = FILL(LINEN_LT if i % 2 else LINEN)
        ws.row_dimensions[r].height = 20
    return ws


def s(key):
    return f"Supuestos!{SUP[key]}"


# ----------------------------------------------------------------------------- RENDER
def render(d, m, out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan Financiero"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN.replace("FF", "", 1)

    build_supuestos(wb, d)   # crea la hoja de inputs y fija el mapa SUP

    widths = {"A": 3, "B": 46, "C": 22, "D": 18, "E": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fase = d.get("fase", "Phase I")

    # --- Banda de título (SIEMPRE arriba, antes de la imagen) --------------------
    def band(r, text, font, fill=GREEN, h=None, align=LEFT):
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]; c.value = text; c.font = font; c.alignment = align
        for col in "BCDE":
            ws[f"{col}{r}"].fill = FILL(fill)
        if h: ws.row_dimensions[r].height = h

    r = 1
    band(r, "L A W A N G", F(26, True, LINEN, "Georgia"), h=46); r += 1
    band(r, f"Lamborghini Wild Coast Sumba  ·  Development Financial Plan  ·  {fase}",
         F(11, False, LINEN, "Georgia", italic=True), h=24); r += 1
    band(r, "OWN · THE · ETERNITY", F(9, True, CANOPY), fill=LAGOON, h=20,
         align=Alignment(horizontal="right", vertical="center", indent=1)); r += 1

    ws.row_dimensions[r].height = 10; r += 1              # aire bajo la banda de título

    # --- Helpers de sección ------------------------------------------------------
    def section(title):
        nonlocal r
        ws.row_dimensions[r].height = 8; r += 1            # aire
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]; c.value = title.upper()
        c.font = F(13, True, LINEN, "Georgia"); c.alignment = LEFT
        for col in "BCDE": ws[f"{col}{r}"].fill = FILL(GREEN)
        ws.row_dimensions[r].height = 26; r += 1

    def colhead(cols):
        nonlocal r
        for col, (txt, al) in cols.items():
            c = ws[f"{col}{r}"]; c.value = txt
            c.font = F(9, True, ASH); c.alignment = al
            c.border = B_BOTTOM
        ws.row_dimensions[r].height = 18; r += 1

    def row_kv(label, value, fmt=None, bold=False, band_i=0, calc=None,
               total=False, indent=True, color=None):
        nonlocal r
        bg = LAGOON if total else (LINEN_LT if band_i % 2 else LINEN)
        txt = LINEN if total else (color or ASH)
        for col in "BCDE":
            cell = ws[f"{col}{r}"]; cell.fill = FILL(bg)
            cell.border = B_TOP_MED if total else B_BOTTOM
        b = ws[f"B{r}"]; b.value = label
        b.font = F(11, bold or total, txt); b.alignment = LEFT if indent else LEFT0
        if calc is not None:
            cc = ws[f"C{r}"]; cc.value = calc
            cc.font = F(9, False, txt if total else SAND); cc.alignment = LEFT0
        v = ws[f"D{r}"]; v.value = value
        v.font = F(11, bold or total, txt); v.alignment = RIGHT
        if fmt: v.number_format = fmt
        ws.row_dimensions[r].height = 20; r += 1

    def pct_cell(value):
        e = ws[f"E{r-1}"]; e.value = value
        e.font = F(10, False, SAND); e.alignment = RIGHT; e.number_format = PCT

    def kpi(label, value, fmt):
        # tarjeta de 2 filas: label (lagoon) + valor grande (lagoon)
        nonlocal r
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]; c.value = label.upper()
        c.font = F(9, True, CANOPY); c.alignment = LEFT
        for col in "BCDE": ws[f"{col}{r}"].fill = FILL(LAGOON)
        ws.row_dimensions[r].height = 18; r += 1
        ws.merge_cells(f"B{r}:E{r}")
        v = ws[f"B{r}"]; v.value = value
        v.font = F(22, True, LINEN, "Georgia"); v.alignment = LEFT
        if fmt: v.number_format = fmt
        for col in "BCDE": ws[f"{col}{r}"].fill = FILL(LAGOON)
        ws.row_dimensions[r].height = 36; r += 1
        ws.row_dimensions[r].height = 6; r += 1            # separación

    # helper para escribir el % (columna E) como fórmula sobre el cell de Ventas
    DENOM = [None]   # se fija tras escribir la línea de Ventas

    def set_pct(row, total=False):
        if not DENOM[0]:
            return
        e = ws[f"E{row}"]; e.value = f"=D{row}/{DENOM[0]}"
        e.number_format = PCT
        e.font = F(10, False, LINEN if total else SAND); e.alignment = RIGHT

    # --- PROJECT SNAPSHOT --------------------------------------------------------
    section(f"Project Snapshot — {fase}")
    row_kv("Nº villas",                 f"={s('num_villas')}",                      None, band_i=0)
    row_kv("Precio de venta por villa", f"={s('precio_villa')}",                     EUR,  band_i=1)
    row_kv("Superficie por villa",      f'={s("superficie_villa")}&" m²"',           None, band_i=2)
    row_kv("Ventas totales (GDV)",      f"={s('num_villas')}*{s('precio_villa')}",   EUR,  band_i=3)
    row_kv("Inicio comercialización",   d.get("inicio_comercializacion", "Mes 2 – 4"), None, band_i=4)
    row_kv("Entrega estimada",          d.get("entrega_estimada", "Mes 12 – 14"),    None, band_i=5)

    # --- ESTADO DE RESULTADOS ----------------------------------------------------
    section(f"Estado de Resultados — {fase}")
    colhead({"B": ("CONCEPTO", LEFT), "C": ("CÁLCULO", LEFT0),
             "D": ("IMPORTE", RIGHT), "E": ("% S/ VENTAS", RIGHT)})

    ventas_row = r
    row_kv("Ventas de villas", f"={s('num_villas')}*{s('precio_villa')}", EUR, band_i=0,
           calc=f'{m["villas"]} × {m["precio"]:,.0f} €'.replace(",", "."))
    DENOM[0] = f"$D${ventas_row}"
    set_pct(ventas_row)

    ti_row = r
    row_kv("TOTAL INGRESOS", f"=D{ventas_row}", EUR, total=True)
    set_pct(ti_row, total=True)

    def cost(lbl, formula, calc="", i=0):
        nonlocal r
        row = r
        row_kv(lbl, formula, EUR, band_i=i, calc=calc)
        set_pct(row)
        return row

    first_cost = r
    cost("Construcción", f"=-{s('num_villas')}*{s('coste_construccion_villa')}",
         f'{m["villas"]} × {d.get("coste_construccion_villa",0):,.0f} €'.replace(",", "."), 0)
    cost("Suelo (adquisición)",            f"=-{s('coste_adquisicion_suelo')}", "", 1)
    cost("Urbanización & infraestructura", f"=-{s('urbanizacion')}",            "", 0)
    cost("Licencias & permisos",           f"=-{s('licencias')}",               "", 1)
    cost("Nick Maltese Studio (diseño & arq.)",
         f"=-({s('num_villas')}*{s('coste_construccion_villa')})*{s('pct_nick_maltese')}",
         "8% s/ construcción", 0)
    cost("Arquitecto local & ingeniería",  f"=-{s('arquitecto_local')}", "", 1)
    cost("Otros profesionales técnicos",   f"=-{s('otros_tecnicos')}",   "", 0)
    cost("Marketing & branding",           f"=-{s('marketing')}",        "", 1)
    cost("Comisiones comerciales",         f"=-{DENOM[0]}*{s('pct_comisiones')}", "% de ventas", 0)
    cost("Gastos legales & notaría",       f"=-{s('legal_notaria')}",    "", 1)
    cost("Gastos administrativos",         f"=-{s('gastos_admin')}",     "", 0)
    last_cost = cost("Seguros & contingencias", f"=-{DENOM[0]}*{s('pct_seguros')}", "% de ventas", 1)

    margen_row = r
    row_kv("MARGEN OPERATIVO ANTES DE REPARTOS",
           f"=D{ventas_row}+SUM(D{first_cost}:D{last_cost})", EUR, total=True)
    set_pct(margen_row, total=True)

    # --- REPARTO DEL BENEFICIO ---------------------------------------------------
    section("Reparto del Beneficio")
    tonino_row = r
    row_kv("Tonino Lamborghini", f"=-D{margen_row}*{s('pct_tonino')}", EUR, band_i=0, calc="3% del beneficio")
    intro_row = r
    row_kv("Introducción de marca", f"=-D{margen_row}*{s('pct_introductor')}", EUR, band_i=1, calc="1% del beneficio")
    neto_row = r
    row_kv("BENEFICIO NETO PARA LAWANG (Fran & Pablo)",
           f"=D{margen_row}+D{tonino_row}+D{intro_row}", EUR, total=True)

    # --- MÉTRICAS CLAVE ----------------------------------------------------------
    section(f"Métricas Clave — {fase}")
    kpi("Ventas (GDV)",               f"=D{ventas_row}", EUR)
    kpi("Beneficio neto Lawang",      f"=D{neto_row}",   EUR)
    kpi("Margen neto sobre ventas",   f"=D{neto_row}/D{ventas_row}", PCT)
    kpi("ROI sobre coste total",      f"=D{neto_row}/-SUM(D{first_cost}:D{last_cost})", PCT)
    kpi("ROI sobre inversión total",
        f"=D{neto_row}/(D{ventas_row}*({s('pct_aportacion_socios')}+{s('pct_deuda')}))", PCT)

    # --- CASH FLOW (si viene en datos) ------------------------------------------
    cf = d.get("cash_flow")
    if cf:
        section(f"Cash Flow Estimado — {fase}")
        colhead({"B": ("HITO", LEFT), "C": ("MES", LEFT0),
                 "D": ("FLUJO", RIGHT), "E": ("ACUMULADO", RIGHT)})
        ws.column_dimensions["H"].hidden = True   # entrada (input)
        ws.column_dimensions["I"].hidden = True   # salida (input)
        prev = None
        for i, h in enumerate(cf):
            row = r
            ws[f"H{row}"] = h.get("entrada", 0) or 0
            ws[f"I{row}"] = h.get("salida", 0) or 0
            for col in "BCDE":
                cell = ws[f"{col}{row}"]; cell.fill = FILL(LINEN_LT if i % 2 else LINEN)
                cell.border = B_BOTTOM
            ws[f"B{row}"].value = h.get("hito", ""); ws[f"B{row}"].font = F(11); ws[f"B{row}"].alignment = LEFT
            ws[f"C{row}"].value = str(h.get("mes", "")); ws[f"C{row}"].font = F(10, False, SAND); ws[f"C{row}"].alignment = LEFT0
            ws[f"D{row}"].value = f"=H{row}-I{row}"; ws[f"D{row}"].number_format = EUR
            ws[f"D{row}"].font = F(11); ws[f"D{row}"].alignment = RIGHT
            ws[f"E{row}"].value = (f"=D{row}" if prev is None else f"=E{prev}+D{row}")
            ws[f"E{row}"].number_format = EUR; ws[f"E{row}"].font = F(11, True, ASH); ws[f"E{row}"].alignment = RIGHT
            ws.row_dimensions[row].height = 20
            prev = row; r += 1

    # --- VALOR DEL SUELO ---------------------------------------------------------
    section("Valor del Suelo")
    coste_row = r
    row_kv("Coste de adquisición", f"={s('coste_adquisicion_suelo')}", EUR, band_i=0,
           calc=f"={s('coste_adquisicion_suelo')}/{s('suelo_total')}&\" €/m²\"")
    pot_row = r
    row_kv("Valor potencial del suelo", f"={s('suelo_total')}*{s('precio_suelo_m2')}", EUR, band_i=1,
           calc=f'{d.get("suelo_total","")} m² × {d.get("precio_suelo_m2","")} €/m²')
    row_kv("Plusvalía estimada", f"=D{pot_row}-D{coste_row}", EUR, total=True)

    # pie
    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    foot = ws[f"B{r}"]
    foot.value = "AxisWorks · Documento de inversión confidencial — Lawang Estate"
    foot.font = F(8, False, SAND, italic=True); foot.alignment = LEFT0

    out = Path(out); out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main(datos_path, salida_path=None):
    with open(datos_path, encoding="utf-8") as f:
        d = json.load(f)
    m = compute(d)
    if salida_path is None:
        nombre = d.get("_salida") or (Path(datos_path).stem + ".xlsx")
        if not nombre.endswith(".xlsx"): nombre += ".xlsx"
        salida_path = BASE / "salida" / nombre
    out = render(d, m, salida_path)
    print(f"OK  Generado: {out}")
    print(f"    Beneficio neto Lawang: {m['net']:,.0f} €  ({m['net_margin']*100:.1f}% s/ventas)")
    print(f"    Margen operativo: {m['op']:,.0f} €   ROI coste: {m['roi_cost']*100:.1f}%   ROI inv: {m['roi_inv']*100:.1f}%")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python build_plan.py datos.json [salida/Plan.xlsx]")
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
